import torch
import torch_scatter
import torch.nn.functional as F
from torch_geometric.nn import MessagePassing
from torch_geometric.nn import Sequential
from torch_geometric.nn import GINConv
from torch.nn import Parameter, Module, Sigmoid,Linear,ReLU,Sequential
from torch_geometric.datasets import TUDataset
from torch_geometric.data import DataLoader
from torch_geometric.nn import global_add_pool
import os.path as osp
import matplotlib.pyplot as plt
from torch_geometric.data import InMemoryDataset
import os
import torch
import scipy.io as sio
import scipy.sparse as sp
import numpy as np
from torch_geometric.data import Data
from openpyxl import load_workbook

wb = load_workbook('AD.xlsx')
sheets = wb.worksheets  # 获取当前所有的sheet

# 获取第一张sheet
sheet1 = sheets[0]
# sheet1 = wb['Sheet']  # 也可以通过已知表名获取sheet
rows = sheet1.rows
columns = sheet1.columns

class YooChooseDataset(InMemoryDataset):
    def __init__(self, root, transform=None, pre_transform=None):
        super(YooChooseDataset, self).__init__(root, transform, pre_transform)
        self.data, self.slices = torch.load(self.processed_paths[0])

    @property
    def raw_file_names(self):
        return []

    @property
    def processed_file_names(self):
        return ['data.pt']

    def download(self):
        pass

    def process(self):

        data_list = []
        filePath = './patientB'
        i = 1
        wb = load_workbook('AD.xlsx')
        sheets = wb.worksheets  # 获取当前所有的sheet

        # 获取第一张sheet
        sheet1 = sheets[0]
        # sheet1 = wb['Sheet']  # 也可以通过已知表名获取sheet
        rows = sheet1.rows
        columns = sheet1.columns
        for root, dirs, files in os.walk(filePath):

            # root 表示当前正在访问的文件夹路径
            # dirs 表示该文件夹下的子目录名list
            # files 表示该文件夹下的文件list

            # 遍历文件
            for f in files:
                for j, k in enumerate(f):
                    if k == "S":
                        rows = sheet1.rows
                        list1 = k + f[j + 1] + f[j + 2] + f[j + 3] + f[j + 4] + f[j + 5] + f[j + 6]
                        # print(list1)
                        for row in rows:
                            # print(2)
                            row_val = [col.value for col in row]
                            if list1 == row_val[0]:
                                # print(row_val[0])
                                if (row_val[3] > 56):
                                    ages = row_val[3] - 56
                                    ages = ages / 40
                                else:
                                    ages = 0
                                if (row_val[2] == "M"):
                                    gender = 1
                                else:
                                    gender = 0
                                break
                Path = os.path.join(root, f)
                load_data = sio.loadmat(Path)
                #print(type(load_data))
                #print(load_data.keys())
                # print(load_data.values())
                # for key, value in load_data.items():
                # print(key, ':', value)
                # print(load_data['ROICorrelation'])
                A = load_data['a']
                edge_index_temp = sp.coo_matrix(A)
                # print(edge_index_temp)
                path2 = 'patient_processed(' + str(i) + ')'
                sp.save_npz(path2, edge_index_temp)
                path3 = './' + path2 + '.npz'
                data = np.load(path3)
                i = i + 1
                # print(data.files)
                # print(data['row'])
                # print(data['col'])
                edge_index = torch.tensor([data['row'], data['col']], dtype=torch.long)
                # print(edge_index)
                #x = np.ones((116, 4)) * 1
                # 116行4列全1矩阵
                data = Data(x=torch.tensor([[0.1, 1.2], [2.2, 3.1], [4.9, 5.2]]), edge_index=edge_index, y=torch.tensor([1], dtype=torch.float),z=torch.tensor([ages], dtype=torch.float),w=torch.tensor([gender], dtype=torch.float))
                data_list.append(data)
                # print(data_list)
        # A = np.array([[0, 1, 0, 1],
        # [1, 0, 1, 0],
        # [0, 1, 0, 0],
        # [1, 0, 0, 0]])
        # print(data_list)
        # print(i)
        filePath = './normalB'
        i = 1
        wb = load_workbook('CN.xlsx')
        sheets = wb.worksheets  # 获取当前所有的sheet

        # 获取第一张sheet
        sheet1 = sheets[0]
        # sheet1 = wb['Sheet']  # 也可以通过已知表名获取sheet
        rows = sheet1.rows
        columns = sheet1.columns
        for root, dirs, files in os.walk(filePath):
            for f in files:
                for j, k in enumerate(f):
                    if k == "S":
                        rows = sheet1.rows
                        list1 = k + f[j + 1] + f[j + 2] + f[j + 3] + f[j + 4] + f[j + 5] + f[j + 6]
                        # print(list1)
                        for row in rows:
                            # print(2)
                            row_val = [col.value for col in row]
                            if list1 == row_val[0]:
                                # print(row_val[0])
                                # ages = row_val[3]
                                if (row_val[3] > 56):
                                    ages = row_val[3] - 56
                                    ages = ages / 40
                                else:
                                    ages = 0
                                if (row_val[2] == "M"):
                                    gender = 1
                                else:
                                    gender = 0
                                break
                Path = os.path.join(root, f)
                load_data = sio.loadmat(Path)
                # print(type(load_data))
                # print(load_data.keys())
                # print(load_data.values())
                # for key, value in load_data.items():
                # print(key, ':', value)
                # print(load_data['ROICorrelation'])
                A = load_data['a']
                edge_index_temp = sp.coo_matrix(A)
                # print(edge_index_temp)
                path2 = '_normal_processed(' + str(i) + ')'
                sp.save_npz(path2, edge_index_temp)
                path3 = './' + path2 + '.npz'
                data = np.load(path3)
                i = i + 1
                # print(data.files)
                # print(data['row'])
                # print(data['col'])
                edge_index = torch.tensor([data['row'], data['col']], dtype=torch.long)
                # print(edge_index)
                x = np.ones((116, 4)) * 0
                # 116行4列全0矩阵
                data = Data(x=torch.tensor([[0.1, 1.2], [2.2, 3.1], [4.9, 5.2]]), edge_index=edge_index, y=torch.tensor([0], dtype=torch.long),z=torch.tensor([ages], dtype=torch.long),w=torch.tensor([gender], dtype=torch.long))
                data_list.append(data)
        # 特征输入

        i = 0
        filePath = './patient_feature'
        for root, dirs, files in os.walk(filePath):
            for f in files:
                Path = os.path.join(root, f)
                tu = open(Path, encoding='gbk')
                txt = []
                for line in tu:
                    list = line.split(',')
                    row_data = [float(x) for x in list]
                    txt.append(row_data)

                data_list[i].x = torch.tensor(txt, dtype=torch.float)
                i = i + 1
        filePath = './normal_feature'
        for root, dirs, files in os.walk(filePath):
            for f in files:
                Path = os.path.join(root, f)
                tu = open(Path, encoding='gbk')
                txt = []
                for line in tu:
                    list = line.split(',')
                    row_data = [float(x) for x in list]
                    txt.append(row_data)

                data_list[i].x = torch.tensor(txt, dtype=torch.float)
                i = i + 1
        data, slices = self.collate(data_list)
        torch.save((data, slices), self.processed_paths[0])

dataset=YooChooseDataset('../')
dataset = dataset.shuffle()
test_dataset = dataset[:272]
train_dataset = dataset[272:]
print(dataset[0].y)  # 1病0正常
test_loader = DataLoader(test_dataset, batch_size=69, shuffle=True)
train_loader = DataLoader(train_dataset, batch_size=69, shuffle=True)
#print(dataset[0].x)
print(train_loader.dataset)
loader = DataLoader(dataset, batch_size=1)
cnt = 0
i = 0
for data in loader:
    i += 1
    if data.y.item() == 1:
        cnt += 1
        
print(cnt, i)