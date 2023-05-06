import torch
import torch_scatter
import torch.nn.functional as F
from torch_geometric.nn import MessagePassing
from torch_geometric.nn import Sequential
from torch_geometric.nn import GINConv
from torch.nn import Parameter, Module, Sigmoid,Linear,ReLU,Sequential
from torch_geometric.datasets import TUDataset
from torch_geometric.data import DataLoader
from torch_geometric.nn import global_add_pool
import os.path as osp
import matplotlib.pyplot as plt
from torch_geometric.data import InMemoryDataset
import os
import torch
import scipy.io as sio
import scipy.sparse as sp
import numpy as np
from torch_geometric.data import Data
from openpyxl import load_workbook
from sklearn.metrics import precision_recall_curve, average_precision_score
wb = load_workbook('AD.xlsx')
sheets = wb.worksheets  # 获取当前所有的sheet
# 获取第一张sheet
sheet1 = sheets[0]
# sheet1 = wb['Sheet']  # 也可以通过已知表名获取sheet
rows = sheet1.rows
columns = sheet1.columns
class YooChooseDataset(InMemoryDataset):
    def __init__(self, root, transform=None, pre_transform=None):
        super(YooChooseDataset, self).__init__(root, transform, pre_transform)
        self.data, self.slices = torch.load(self.processed_paths[0])

    @property
    def raw_file_names(self):
        return []

    @property
    def processed_file_names(self):
        return ['data.pt']

    def download(self):
        pass

    def process(self):
        data_list = []
        filePath = './normalB'
        i = 1
        wb = load_workbook('CN.xlsx')
        sheets = wb.worksheets  # 获取当前所有的sheet

        # 获取第一张sheet
        sheet1 = sheets[0]
        # sheet1 = wb['Sheet']  # 也可以通过已知表名获取sheet
        rows = sheet1.rows
        columns = sheet1.columns
        for root, dirs, files in os.walk(filePath):
            for f in files:
                for j, k in enumerate(f):
                    if k == "S":
                        rows = sheet1.rows
                        list1 = k + f[j + 1] + f[j + 2] + f[j + 3] + f[j + 4] + f[j + 5] + f[j + 6]
                        # print(list1)
                        for row in rows:
                            # print(2)
                            row_val = [col.value for col in row]
                            if list1 == row_val[0]:
                                # print(row_val[0])
                                # ages = row_val[3]
                                if (row_val[3] > 56):
                                    ages = row_val[3] - 56
                                    ages = ages / 40
                                else:
                                    ages = 0
                                if (row_val[2] == "M"):
                                    gender = 1
                                else:
                                    gender = 0
                                break
                Path = os.path.join(root, f)
                load_data = sio.loadmat(Path)
                # print(type(load_data))
                # print(load_data.keys())
                # print(load_data.values())
                # for key, value in load_data.items():
                # print(key, ':', value)
                # print(load_data['ROICorrelation'])
                A = load_data['a']
                edge_index_temp = sp.coo_matrix(A)
                # print(edge_index_temp)
                path2 = '_normal_processed(' + str(i) + ')'
                sp.save_npz(path2, edge_index_temp)
                path3 = './' + path2 + '.npz'
                data = np.load(path3)
                i = i + 1
                # print(data.files)
                # print(data['row'])
                # print(data['col'])
                edge_index = torch.tensor([data['row'], data['col']], dtype=torch.long)
                # print(edge_index)
                x = np.ones((116, 4)) * 0
                # 116行4列全0矩阵
                data = Data(x=torch.tensor([[0.1, 1.2], [2.2, 3.1], [4.9, 5.2]]), edge_index=edge_index, y=torch.tensor([0], dtype=torch.long))
                data_list.append(data)
        # 特征输入
        i = 0
        filePath = './normal_feature'
        for root, dirs, files in os.walk(filePath):
            for f in files:
                Path = os.path.join(root, f)
                tu = open(Path, encoding='gbk')
                txt = []
                for line in tu:
                    list = line.split(',')
                    row_data = [float(x) for x in list]
                    txt.append(row_data)

                data_list[i].x = torch.tensor(txt, dtype=torch.float)
                i = i + 1
        data, slices = self.collate(data_list)
        torch.save((data, slices), self.processed_paths[0])

class YooChooseDataset1(InMemoryDataset):
    def __init__(self, root, transform=None, pre_transform=None):
        super(YooChooseDataset1, self).__init__(root, transform, pre_transform)
        self.data, self.slices = torch.load(self.processed_paths[0])

    @property
    def raw_file_names(self):
        return []

    @property
    def processed_file_names(self):
        return ['data.pt']

    def download(self):
        pass

    def process(self):
        data_list = []
        filePath = './patientB'
        i = 1
        wb = load_workbook('CN.xlsx')
        sheets = wb.worksheets  # 获取当前所有的sheet

        # 获取第一张sheet
        sheet1 = sheets[0]
        # sheet1 = wb['Sheet']  # 也可以通过已知表名获取sheet
        rows = sheet1.rows
        columns = sheet1.columns
        for root, dirs, files in os.walk(filePath):
            for f in files:
                for j, k in enumerate(f):
                    if k == "S":
                        rows = sheet1.rows
                        list1 = k + f[j + 1] + f[j + 2] + f[j + 3] + f[j + 4] + f[j + 5] + f[j + 6]
                        # print(list1)
                        for row in rows:
                            # print(2)
                            row_val = [col.value for col in row]
                            if list1 == row_val[0]:
                                # print(row_val[0])
                                # ages = row_val[3]
                                if (row_val[3] > 56):
                                    ages = row_val[3] - 56
                                    ages = ages / 40
                                else:
                                    ages = 0
                                if (row_val[2] == "M"):
                                    gender = 1
                                else:
                                    gender = 0
                                break
                Path = os.path.join(root, f)
                load_data = sio.loadmat(Path)
                # print(type(load_data))
                # print(load_data.keys())
                # print(load_data.values())
                # for key, value in load_data.items():
                # print(key, ':', value)
                # print(load_data['ROICorrelation'])
                A = load_data['a']
                edge_index_temp = sp.coo_matrix(A)
                # print(edge_index_temp)
                path2 = '_patient_processed(' + str(i) + ')'
                sp.save_npz(path2, edge_index_temp)
                path3 = './' + path2 + '.npz'
                data = np.load(path3)
                i = i + 1
                # print(data.files)
                # print(data['row'])
                # print(data['col'])
                edge_index = torch.tensor([data['row'], data['col']], dtype=torch.long)
                # print(edge_index)
                x = np.ones((116, 4)) * 0
                # 116行4列全0矩阵
                data = Data(x=torch.tensor([[0.1, 1.2], [2.2, 3.1], [4.9, 5.2]]), edge_index=edge_index, y=torch.tensor([0], dtype=torch.long))
                data_list.append(data)
        # 特征输入
        i = 0
        filePath = './patient_feature'
        for root, dirs, files in os.walk(filePath):
            for f in files:
                Path = os.path.join(root, f)
                tu = open(Path, encoding='gbk')
                txt = []
                for line in tu:
                    list = line.split(',')
                    row_data = [float(x) for x in list]
                    txt.append(row_data)

                data_list[i].x = torch.tensor(txt, dtype=torch.float)
                i = i + 1
        data, slices = self.collate(data_list)
        torch.save((data, slices), self.processed_paths[0])

class GINNet(torch.nn.Module):
    def __init__(self):
        super(GINNet, self).__init__()
        #torch.manual_seed(420)
        num_features = 6
        dim = 32
        # assign the weights for each task
        self.weights = torch.nn.Parameter(torch.ones(3).float(), requires_grad=True)

        nn1 = Sequential(Linear(num_features, dim), ReLU(), Linear(dim, dim))
        self.conv1 = GINConv(nn1)
        self.bn1 = torch.nn.BatchNorm1d(dim)

        nn2 = Sequential(Linear(dim, dim), ReLU(), Linear(dim, dim))
        self.conv2 = GINConv(nn2)
        self.bn2 = torch.nn.BatchNorm1d(dim)

        nn3 = Sequential(Linear(dim, dim), ReLU(), Linear(dim, dim))
        self.conv3 = GINConv(nn3)
        self.bn3 = torch.nn.BatchNorm1d(dim)

        nn4 = Sequential(Linear(dim, dim), ReLU(), Linear(dim, dim))
        self.conv4 = GINConv(nn4)
        self.bn4 = torch.nn.BatchNorm1d(dim)

        nn5 = Sequential(Linear(dim, dim), ReLU(), Linear(dim, dim))
        self.conv5 = GINConv(nn5)
        self.bn5 = torch.nn.BatchNorm1d(dim)

        self.fc1 = Linear(dim, dim)
        self.fc2 = Linear(dim, 2)
        self.age_fc_layers = Sequential(Linear(dim, dim), ReLU(), Linear(dim, 1), Sigmoid())
        self.gender_fc_layers = Sequential(Linear(dim, dim), ReLU(), Linear(dim, 2))



    def forward(self, x, edge_index, batch):
        x = F.relu(self.conv1(x, edge_index))
        x = self.bn1(x)
        x = F.relu(self.conv2(x, edge_index))
        x = self.bn2(x)
        # x = F.relu(self.conv3(x, edge_index))
        # x = self.bn3(x)
        # x = F.relu(self.conv4(x, edge_index))
        # x = self.bn4(x)
        # x = F.relu(self.conv5(x, edge_index))
        # x = self.bn5(x)
        x = global_add_pool(x, batch)
        x = F.relu(self.fc1(x))
        x = F.dropout(x, p=0.5, training=self.training)
        x0 = self.fc2(x)
        x1 = self.gender_fc_layers(x)
        x2 = self.age_fc_layers(x)
        return [F.log_softmax(x0, dim=-1), F.log_softmax(x1, dim=-1), F.log_softmax(x2, dim=-1)]


class predict:
    def __init__(self):
        self.dataset = YooChooseDataset1('../')
        self.dataset.process()
        print("Successfully create!")
        self.test_dataset = self.dataset
        self.test_loader = DataLoader(self.test_dataset, batch_size=1, shuffle=True)
        self.dataset = YooChooseDataset('../')
        self.dataset.process()
        print("Successfully create!")
        self.train_dataset = self.dataset
        self.train_loader = DataLoader(self.train_dataset, batch_size=1, shuffle=True)

    def work(self):
        device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
        model = GINNet().to(device)
        model.eval()
        model.load_state_dict(torch.load('model_params.pt'))
        # print(len(self.test_loader.dataset))
        cnt = 0
        y_true = torch.ones(1360)
        y_true[:770] = 0
        y_scores = torch.zeros(1360)
        tmp = 0
        for data in self.test_loader:
            data = data.to(device)
            print(data.y)
            if data.y.item() == 1:
                tmp -= 1
            output = model(data.x, data.edge_index, data.batch)
            # print(output)
            pred = output[0].max(dim=1)[1]
            if pred[0].item() == 1:
                y_scores[cnt] = 1
                # print("patient here")
            cnt += 1
        print(tmp)
        print(cnt)
        tmp2 = 0
        for data in self.train_loader:
            data = data.to(device)
            
            if data.y == 1:
                tmp2 -= 1
            output = model(data.x, data.edge_index, data.batch)
            # print(output)
            pred = output[0].max(dim=1)[1]
            if pred[0].item() == 1:
                y_scores[cnt] = 1
                # print("normal here")
            cnt += 1
        print(tmp2)
        print(cnt)
        precision, recall, thresholds = precision_recall_curve(y_true, y_scores)
        average_precision = average_precision_score(y_true, y_scores)
        # print(y_scores[:40])
        # 画PR曲线
        plt.plot(recall, precision, label='AP={:.3f}'.format(average_precision))
        plt.xlabel('Recall')
        plt.ylabel('Precision')
        plt.legend(loc='best')
        plt.title('Precision-Recall curve')
        plt.show()
a = predict()
a.work()