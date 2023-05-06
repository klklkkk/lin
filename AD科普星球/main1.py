import torch
import torch_scatter
import torch.nn.functional as F
from torch_geometric.nn import MessagePassing
from torch_geometric.nn import Sequential
from torch_geometric.nn import GINConv
from torch.nn import Parameter, Module, Sigmoid,Linear,ReLU,Sequential
from torch_geometric.datasets import TUDataset
from torch_geometric.data import DataLoader
from torch_geometric.nn import global_add_pool
import os.path as osp
import matplotlib.pyplot as plt
from torch_geometric.data import InMemoryDataset
import os
import torch
import scipy.io as sio
import scipy.sparse as sp
import numpy as np
from torch_geometric.data import Data
from openpyxl import load_workbook

wb = load_workbook('AD.xlsx')
sheets = wb.worksheets  # 获取当前所有的sheet

# 获取第一张sheet
sheet1 = sheets[0]
# sheet1 = wb['Sheet']  # 也可以通过已知表名获取sheet
rows = sheet1.rows
columns = sheet1.columns

class YooChooseDataset(InMemoryDataset):
    def __init__(self, root, transform=None, pre_transform=None):
        super(YooChooseDataset, self).__init__(root, transform, pre_transform)
        self.data, self.slices = torch.load(self.processed_paths[0])

    @property
    def raw_file_names(self):
        return []

    @property
    def processed_file_names(self):
        return ['data.pt']

    def download(self):
        pass

    def process(self):

        data_list = []
        filePath = './patientB'
        i = 1
        wb = load_workbook('AD.xlsx')
        sheets = wb.worksheets  # 获取当前所有的sheet

        # 获取第一张sheet
        sheet1 = sheets[0]
        # sheet1 = wb['Sheet']  # 也可以通过已知表名获取sheet
        rows = sheet1.rows
        columns = sheet1.columns
        for root, dirs, files in os.walk(filePath):

            # root 表示当前正在访问的文件夹路径
            # dirs 表示该文件夹下的子目录名list
            # files 表示该文件夹下的文件list

            # 遍历文件
            for f in files:
                for j, k in enumerate(f):
                    if k == "S":
                        rows = sheet1.rows
                        list1 = k + f[j + 1] + f[j + 2] + f[j + 3] + f[j + 4] + f[j + 5] + f[j + 6]
                        # print(list1)
                        for row in rows:
                            # print(2)
                            row_val = [col.value for col in row]
                            if list1 == row_val[0]:
                                # print(row_val[0])
                                if (row_val[3] > 56):
                                    ages = row_val[3] - 56
                                    ages = ages / 40
                                else:
                                    ages = 0
                                if (row_val[2] == "M"):
                                    gender = 1
                                else:
                                    gender = 0
                                break
                Path = os.path.join(root, f)
                load_data = sio.loadmat(Path)
                #print(type(load_data))
                #print(load_data.keys())
                # print(load_data.values())
                # for key, value in load_data.items():
                # print(key, ':', value)
                # print(load_data['ROICorrelation'])
                A = load_data['a']
                edge_index_temp = sp.coo_matrix(A)
                # print(edge_index_temp)
                path2 = 'patient_processed(' + str(i) + ')'
                sp.save_npz(path2, edge_index_temp)
                path3 = './' + path2 + '.npz'
                data = np.load(path3)
                i = i + 1
                # print(data.files)
                # print(data['row'])
                # print(data['col'])
                edge_index = torch.tensor([data['row'], data['col']], dtype=torch.long)
                # print(edge_index)
                #x = np.ones((116, 4)) * 1
                # 116行4列全1矩阵
                data = Data(x=torch.tensor([[0.1, 1.2], [2.2, 3.1], [4.9, 5.2]]), edge_index=edge_index, y=torch.tensor([1], dtype=torch.float),z=torch.tensor([ages], dtype=torch.float),w=torch.tensor([gender], dtype=torch.float))
                data_list.append(data)
                # print(data_list)
        # A = np.array([[0, 1, 0, 1],
        # [1, 0, 1, 0],
        # [0, 1, 0, 0],
        # [1, 0, 0, 0]])
        # print(data_list)
        # print(i)
        filePath = './normalB'
        i = 1
        wb = load_workbook('CN.xlsx')
        sheets = wb.worksheets  # 获取当前所有的sheet

        # 获取第一张sheet
        sheet1 = sheets[0]
        # sheet1 = wb['Sheet']  # 也可以通过已知表名获取sheet
        rows = sheet1.rows
        columns = sheet1.columns
        for root, dirs, files in os.walk(filePath):
            for f in files:
                for j, k in enumerate(f):
                    if k == "S":
                        rows = sheet1.rows
                        list1 = k + f[j + 1] + f[j + 2] + f[j + 3] + f[j + 4] + f[j + 5] + f[j + 6]
                        # print(list1)
                        for row in rows:
                            # print(2)
                            row_val = [col.value for col in row]
                            if list1 == row_val[0]:
                                # print(row_val[0])
                                # ages = row_val[3]
                                if (row_val[3] > 56):
                                    ages = row_val[3] - 56
                                    ages = ages / 40
                                else:
                                    ages = 0
                                if (row_val[2] == "M"):
                                    gender = 1
                                else:
                                    gender = 0
                                break
                Path = os.path.join(root, f)
                load_data = sio.loadmat(Path)
                # print(type(load_data))
                # print(load_data.keys())
                # print(load_data.values())
                # for key, value in load_data.items():
                # print(key, ':', value)
                # print(load_data['ROICorrelation'])
                A = load_data['a']
                edge_index_temp = sp.coo_matrix(A)
                # print(edge_index_temp)
                path2 = '_normal_processed(' + str(i) + ')'
                sp.save_npz(path2, edge_index_temp)
                path3 = './' + path2 + '.npz'
                data = np.load(path3)
                i = i + 1
                # print(data.files)
                # print(data['row'])
                # print(data['col'])
                edge_index = torch.tensor([data['row'], data['col']], dtype=torch.long)
                # print(edge_index)
                x = np.ones((116, 4)) * 0
                # 116行4列全0矩阵
                data = Data(x=torch.tensor([[0.1, 1.2], [2.2, 3.1], [4.9, 5.2]]), edge_index=edge_index, y=torch.tensor([0], dtype=torch.long),z=torch.tensor([ages], dtype=torch.long),w=torch.tensor([gender], dtype=torch.long))
                data_list.append(data)
        # 特征输入

        i = 0
        filePath = './patient_feature'
        for root, dirs, files in os.walk(filePath):
            for f in files:
                Path = os.path.join(root, f)
                tu = open(Path, encoding='gbk')
                txt = []
                for line in tu:
                    list = line.split(',')
                    row_data = [float(x) for x in list]
                    txt.append(row_data)

                data_list[i].x = torch.tensor(txt, dtype=torch.float)
                i = i + 1
        filePath = './normal_feature'
        for root, dirs, files in os.walk(filePath):
            for f in files:
                Path = os.path.join(root, f)
                tu = open(Path, encoding='gbk')
                txt = []
                for line in tu:
                    list = line.split(',')
                    row_data = [float(x) for x in list]
                    txt.append(row_data)

                data_list[i].x = torch.tensor(txt, dtype=torch.float)
                i = i + 1
        data, slices = self.collate(data_list)
        torch.save((data, slices), self.processed_paths[0])

dataset=YooChooseDataset('../')
dataset = dataset.shuffle()
test_dataset = dataset[:272]
train_dataset = dataset[272:]
print(dataset[0].y)  # 1病0正常
print(dataset[-1].y)
print("----")
# print(dataset[0].w)  # 1男0女
# print(dataset[0].z)  # 年龄
test_loader = DataLoader(test_dataset, batch_size=69, shuffle=True)
train_loader = DataLoader(train_dataset, batch_size=69, shuffle=True)
#print(dataset[0].x)
print(train_loader.dataset)
class GINNet(torch.nn.Module):
    def __init__(self):
        super(GINNet, self).__init__()
        #torch.manual_seed(420)
        num_features = dataset.num_features
        dim = 32
        # assign the weights for each task
        self.weights = torch.nn.Parameter(torch.ones(3).float(), requires_grad=True)

        nn1 = Sequential(Linear(num_features, dim), ReLU(), Linear(dim, dim))
        self.conv1 = GINConv(nn1)
        self.bn1 = torch.nn.BatchNorm1d(dim)

        nn2 = Sequential(Linear(dim, dim), ReLU(), Linear(dim, dim))
        self.conv2 = GINConv(nn2)
        self.bn2 = torch.nn.BatchNorm1d(dim)

        nn3 = Sequential(Linear(dim, dim), ReLU(), Linear(dim, dim))
        self.conv3 = GINConv(nn3)
        self.bn3 = torch.nn.BatchNorm1d(dim)

        nn4 = Sequential(Linear(dim, dim), ReLU(), Linear(dim, dim))
        self.conv4 = GINConv(nn4)
        self.bn4 = torch.nn.BatchNorm1d(dim)

        nn5 = Sequential(Linear(dim, dim), ReLU(), Linear(dim, dim))
        self.conv5 = GINConv(nn5)
        self.bn5 = torch.nn.BatchNorm1d(dim)

        self.fc1 = Linear(dim, dim)
        self.fc2 = Linear(dim, dataset.num_classes)
        self.age_fc_layers = Sequential(Linear(dim, dim), ReLU(), Linear(dim, 1), Sigmoid())
        self.gender_fc_layers = Sequential(Linear(dim, dim), ReLU(), Linear(dim, 2))



    def forward(self, x, edge_index, batch):
        x = F.relu(self.conv1(x, edge_index))
        x = self.bn1(x)
        x = F.relu(self.conv2(x, edge_index))
        x = self.bn2(x)
        # x = F.relu(self.conv3(x, edge_index))
        # x = self.bn3(x)
        # x = F.relu(self.conv4(x, edge_index))
        # x = self.bn4(x)
        # x = F.relu(self.conv5(x, edge_index))
        # x = self.bn5(x)
        x = global_add_pool(x, batch)
        x = F.relu(self.fc1(x))
        x = F.dropout(x, p=0.5, training=self.training)
        x0 = self.fc2(x)
        x1 = self.gender_fc_layers(x)
        x2 = self.age_fc_layers(x)
        return [F.log_softmax(x0, dim=-1), F.log_softmax(x1, dim=-1), F.log_softmax(x2, dim=-1)]


device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
Loss_list = []
Accuracy_list = []
model =GINNet().to(device)
mse = torch.nn.MSELoss().to(device)  # 损失函数选用均方误差损失或其它损失
optimizer = torch.optim.Adam(model.parameters(), lr=0.0045)


def train(epoch):
    model.train()
    if epoch == 101:
        for param_group in optimizer.param_groups:
            param_group['lr'] = 0.5 * param_group['lr']
    if epoch == 201:
        for param_group in optimizer.param_groups:
            param_group['lr'] = 0.5 * param_group['lr']
    if epoch == 301:
        for param_group in optimizer.param_groups:
            param_group['lr'] = 0.5 * param_group['lr']
    if epoch == 401:
        for param_group in optimizer.param_groups:
            param_group['lr'] = 0.5 * param_group['lr']
    if epoch == 501:
        for param_group in optimizer.param_groups:
            param_group['lr'] = 0.5 * param_group['lr']

    loss_all = 0
    for data in train_loader:
        data = data.to(device)
        optimizer.zero_grad()
        output = model(data.x, data.edge_index, data.batch)
        loss0 = F.nll_loss(output[0], data.y.long())
        loss1 = F.nll_loss(output[1], data.w.long())
        loss2 = mse(output[2], data.z)
        task_loss = torch.stack([loss0, loss1, loss2], dim=-1)
        weighted_task_loss = torch.mul(model.weights, task_loss)
        if epoch == 1:
            # set L(0)
            global initial_task_loss
            if torch.cuda.is_available():
                initial_task_loss = task_loss.data.cpu()
            else:
                initial_task_loss = task_loss.data
            initial_task_loss = initial_task_loss.numpy()

        # loss = (loss0 + loss1 + loss2) / 3
        loss = torch.sum(weighted_task_loss)
        optimizer.zero_grad()
        loss.backward(retain_graph=True)

        model.weights.grad.data = model.weights.grad.data * 0.0

        # grandNorm
        alpha = 0.3
        # get layer of shared weights
        W = model.fc1
        # get the gradient norms for each of the tasks
        # G^{(i)}_w(t)
        norms = []
        for i in range(len(task_loss)):
            # get the gradient of this task loss with respect to the shared parameters
            gygw = torch.autograd.grad(task_loss[i], W.parameters(), retain_graph=True)
            # compute the norm
            norms.append(torch.norm(torch.mul(model.weights[i], gygw[0])))
        norms = torch.stack(norms)
        # print('G_w(t): {}'.format(norms))

        # compute the inverse training rate r_i(t)
        # \curl{L}_i
        if torch.cuda.is_available():
            loss_ratio = task_loss.data.cpu().numpy() / initial_task_loss
        else:
            loss_ratio = task_loss.data.numpy() / initial_task_loss
        # r_i(t)
        inverse_train_rate = loss_ratio / np.mean(loss_ratio)
        # print('r_i(t): {}'.format(inverse_train_rate))

        # compute the mean norm \tilde{G}_w(t)
        if torch.cuda.is_available():
            mean_norm = np.mean(norms.data.cpu().numpy())
        else:
            mean_norm = np.mean(norms.data.numpy())
        # print('tilde G_w(t): {}'.format(mean_norm))

        # compute the GradNorm loss
        # this term has to remain constant
        constant_term = torch.tensor(mean_norm * (inverse_train_rate ** alpha), requires_grad=False)
        if torch.cuda.is_available():
            constant_term = constant_term.cuda()
        # print('Constant term: {}'.format(constant_term))
        # this is the GradNorm loss itself
        grad_norm_loss = torch.tensor(torch.sum(torch.abs(norms - constant_term)), requires_grad=True)
        # print('GradNorm loss {}'.format(grad_norm_loss))

        # compute the gradient for the weights
        model.weights.grad = torch.autograd.grad(grad_norm_loss, model.weights, allow_unused=True)[0]


        loss_all += loss.item() * data.num_graphs
        optimizer.step()
    return loss_all / len(train_dataset)


def test(loader):
    model.eval()

    correct = 0
    for data in loader:
        data = data.to(device)
        output = model(data.x, data.edge_index, data.batch)
        pred = output[0].max(dim=1)[1]
        correct += pred.eq(data.y).sum().item()
    return correct / len(loader.dataset)


for epoch in range(1, 601):
    train_loss = train(epoch)
    train_acc = test(train_loader)
    test_acc = test(test_loader)
    print('Epoch: {:03d}, Train Loss: {:.7f}, '
          'Train Acc: {:.7f}, Test Acc: {:.7f}'.format(epoch, train_loss,
                                                       train_acc, test_acc))
    Loss_list.append(train_acc)
    Accuracy_list.append(test_acc)

x1 = range(0, 600)
x2 = range(0, 600)
y1 = Accuracy_list
y2 = Loss_list
plt.subplot(2, 1, 1)
plt.plot(x1, y1, 'o-')
plt.title('Test accuracy vs. epoches')
plt.ylabel('test_acc')
plt.subplot(2, 1, 2)
plt.plot(x2, y2, '.-')
plt.xlabel('train_acc vs. epoches')
plt.ylabel('train_acc')
plt.show()